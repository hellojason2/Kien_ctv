"""
Excel Export Module
Utilities for generating XLSX Excel files from data.

# ══════════════════════════════════════════════════════════════════════════════
# MODULE STRUCTURE MAP
# ══════════════════════════════════════════════════════════════════════════════
#
# FUNCTIONS:
# - create_xlsx_response(data, columns, filename, sheet_name) -> Flask Response
#   DOES: Creates an Excel file from data and returns as downloadable response
#   INPUTS: data (list of dicts), columns (list of column configs), filename, sheet_name
#   OUTPUTS: Flask Response with Excel file attachment
#
# - style_header_row(ws, num_cols) -> None
#   DOES: Applies styling to header row (bold, background color)
#   INPUTS: worksheet, number of columns
#
# - auto_size_columns(ws) -> None
#   DOES: Auto-adjusts column widths based on content
#   INPUTS: worksheet
#
# - format_currency(value) -> str
#   DOES: Formats number as VND currency
#   INPUTS: numeric value
#   OUTPUTS: Formatted string
#
# ══════════════════════════════════════════════════════════════════════════════

Created: December 30, 2025
"""

import io
from datetime import datetime
from flask import Response
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def create_xlsx_response(data, columns, filename, sheet_name="Data"):
    """
    Create an Excel file from data and return as Flask Response
    
    DOES: Generates a styled XLSX file with headers and data
    INPUTS:
        - data: List of dictionaries containing the data
        - columns: List of column configurations, each with:
            - 'key': Key in data dict
            - 'header': Display header text
            - 'width': Optional column width (default: auto)
            - 'format': Optional format type ('currency', 'date', 'percent')
        - filename: Name of the file (without extension)
        - sheet_name: Name of the Excel sheet
    OUTPUTS: Flask Response with Excel file as attachment
    
    Example:
        columns = [
            {'key': 'ma_ctv', 'header': 'CTV Code', 'width': 15},
            {'key': 'ten', 'header': 'Name', 'width': 25},
            {'key': 'commission', 'header': 'Commission', 'format': 'currency'}
        ]
        return create_xlsx_response(data, columns, 'ctv_export')
    """
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write headers
    for col_idx, col_config in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_config['header'])
    
    # Style headers
    style_header_row(ws, len(columns))
    
    # Write data rows
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, col_config in enumerate(columns, 1):
            key = col_config['key']
            value = row_data.get(key, '')
            
            # Handle None values
            if value is None:
                value = ''
            
            # Format values based on type
            format_type = col_config.get('format')
            if format_type == 'currency' and value != '':
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            elif format_type == 'percent' and value != '':
                try:
                    value = float(value)
                except (ValueError, TypeError):
                    pass
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Apply number format
            if format_type == 'currency':
                cell.number_format = '#,##0'
            elif format_type == 'percent':
                cell.number_format = '0.00%'
            
            # Center align certain columns
            if format_type in ('currency', 'percent'):
                cell.alignment = Alignment(horizontal='right')
    
    # Set column widths
    for col_idx, col_config in enumerate(columns, 1):
        if 'width' in col_config:
            ws.column_dimensions[get_column_letter(col_idx)].width = col_config['width']
    
    # Auto-size columns that don't have explicit width
    auto_size_columns(ws, columns)
    
    # Add borders to all cells
    add_borders(ws, len(data) + 1, len(columns))
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save to bytes buffer
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    full_filename = f"{filename}_{timestamp}.xlsx"
    
    return Response(
        output.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={
            'Content-Disposition': f'attachment; filename={full_filename}'
        }
    )


def style_header_row(ws, num_cols):
    """
    Apply styling to the header row
    
    DOES: Makes headers bold with colored background
    INPUTS: worksheet, number of columns
    """
    header_fill = PatternFill(start_color='1a1a24', end_color='1a1a24', fill_type='solid')
    header_font = Font(bold=True, color='06b6d4', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment


def auto_size_columns(ws, columns):
    """
    Auto-adjust column widths based on content
    
    DOES: Calculates optimal width for each column
    INPUTS: worksheet, column configurations
    """
    for col_idx, col_config in enumerate(columns, 1):
        if 'width' in col_config:
            continue  # Skip if explicit width is set
        
        max_length = len(str(col_config['header']))
        column_letter = get_column_letter(col_idx)
        
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=col_idx, max_col=col_idx):
            for cell in row:
                try:
                    cell_length = len(str(cell.value)) if cell.value else 0
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
        
        # Add some padding
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def add_borders(ws, num_rows, num_cols):
    """
    Add borders to all data cells
    
    DOES: Applies thin borders to create grid effect
    INPUTS: worksheet, number of rows, number of columns
    """
    thin_border = Border(
        left=Side(style='thin', color='27272a'),
        right=Side(style='thin', color='27272a'),
        top=Side(style='thin', color='27272a'),
        bottom=Side(style='thin', color='27272a')
    )
    
    for row in range(1, num_rows + 1):
        for col in range(1, num_cols + 1):
            ws.cell(row=row, column=col).border = thin_border


def format_currency_vnd(value):
    """
    Format a number as VND currency string
    
    DOES: Converts number to formatted VND string
    INPUTS: numeric value
    OUTPUTS: Formatted string (e.g., "1,500,000")
    """
    if value is None:
        return ''
    try:
        return f"{float(value):,.0f}"
    except (ValueError, TypeError):
        return str(value)


# ══════════════════════════════════════════════════════════════════════════════
# COLUMN CONFIGURATIONS FOR DIFFERENT EXPORTS
# ══════════════════════════════════════════════════════════════════════════════

CTV_EXPORT_COLUMNS = [
    {'key': 'ma_ctv', 'header': 'Ma CTV', 'width': 12},
    {'key': 'ten', 'header': 'Ho Ten', 'width': 25},
    {'key': 'sdt', 'header': 'SDT', 'width': 15},
    {'key': 'email', 'header': 'Email', 'width': 30},
    {'key': 'nguoi_gioi_thieu', 'header': 'Nguoi Gioi Thieu', 'width': 15},
    {'key': 'nguoi_gioi_thieu_name', 'header': 'Ten NGT', 'width': 25},
    {'key': 'cap_bac', 'header': 'Cap Bac', 'width': 12},
    {'key': 'is_active', 'header': 'Trang Thai', 'width': 12},
    {'key': 'created_at', 'header': 'Ngay Tao', 'width': 20},
]

COMMISSION_EXPORT_COLUMNS = [
    {'key': 'id', 'header': 'ID', 'width': 8},
    {'key': 'transaction_id', 'header': 'Ma GD', 'width': 10},
    {'key': 'ctv_code', 'header': 'Ma CTV', 'width': 12},
    {'key': 'ctv_name', 'header': 'Ten CTV', 'width': 25},
    {'key': 'level', 'header': 'Cap', 'width': 6},
    {'key': 'commission_rate', 'header': 'Ti Le', 'width': 10, 'format': 'percent'},
    {'key': 'transaction_amount', 'header': 'Gia Tri GD', 'width': 15, 'format': 'currency'},
    {'key': 'commission_amount', 'header': 'Hoa Hong', 'width': 15, 'format': 'currency'},
    {'key': 'created_at', 'header': 'Ngay Tao', 'width': 20},
]

COMMISSION_SUMMARY_COLUMNS = [
    {'key': 'ctv_code', 'header': 'Ma CTV', 'width': 12},
    {'key': 'ctv_name', 'header': 'Ten CTV', 'width': 25},
    {'key': 'ctv_phone', 'header': 'SDT', 'width': 15},
    {'key': 'total_service_price', 'header': 'Tong Doanh Thu', 'width': 18, 'format': 'currency'},
    {'key': 'total_commission', 'header': 'Tong Hoa Hong', 'width': 18, 'format': 'currency'},
]

CLIENTS_EXPORT_COLUMNS = [
    {'key': 'ten_khach', 'header': 'Ten Khach', 'width': 25},
    {'key': 'sdt', 'header': 'SDT', 'width': 15},
    {'key': 'co_so', 'header': 'Co So', 'width': 20},
    {'key': 'nguoi_chot', 'header': 'Nguoi Chot', 'width': 12},
    {'key': 'service_count', 'header': 'So DV', 'width': 10},
    {'key': 'first_visit_date', 'header': 'Lan Dau', 'width': 12},
    {'key': 'overall_status', 'header': 'Trang Thai', 'width': 15},
    {'key': 'overall_deposit', 'header': 'Coc', 'width': 12},
]

ACTIVITY_LOG_COLUMNS = [
    {'key': 'id', 'header': 'ID', 'width': 8},
    {'key': 'timestamp', 'header': 'Thoi Gian', 'width': 20},
    {'key': 'event_type', 'header': 'Loai Su Kien', 'width': 18},
    {'key': 'user_type', 'header': 'Loai User', 'width': 12},
    {'key': 'user_id', 'header': 'User ID', 'width': 15},
    {'key': 'ip_address', 'header': 'IP', 'width': 15},
    {'key': 'endpoint', 'header': 'Endpoint', 'width': 30},
    {'key': 'method', 'header': 'Method', 'width': 8},
    {'key': 'status_code', 'header': 'Status', 'width': 8},
    {'key': 'details', 'header': 'Chi Tiet', 'width': 40},
]

COMMISSION_SETTINGS_COLUMNS = [
    {'key': 'level', 'header': 'Cap', 'width': 8},
    {'key': 'rate', 'header': 'Ti Le', 'width': 12, 'format': 'percent'},
    {'key': 'description', 'header': 'Mo Ta', 'width': 40},
    {'key': 'updated_at', 'header': 'Cap Nhat', 'width': 20},
    {'key': 'updated_by', 'header': 'Nguoi Cap Nhat', 'width': 15},
]

